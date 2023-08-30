import os
import psycopg2
from openpyxl import load_workbook

db_params = {
    'dbname': 'db1',
    'user': 'enid',
    'password': 'enid',
    'host': 'localhost',
}

excel_file_path = os.path.join('table-data.xlsx')

create_table_query = """
CREATE TABLE IF NOT EXISTS TableData (
    Date DATE, 
    Chart_Type TEXT, 
    Customer_Id INT, 
    Customer_Name TEXT, 
    Days_Overdue INT, 
    Amount_Outstanding INT, 
    Recovery INT
);
"""

conn = psycopg2.connect(**db_params) # connection
cursor = conn.cursor()

cursor.execute(create_table_query) # table creation
conn.commit()

wb = load_workbook(excel_file_path)
sheet = wb.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    Date, Chart_Type, Customer_Id, Customer_Name, Days_Overdue, Amount_Outstanding, Recovery = row

    # empty cells
    if Recovery is None:
        Recovery = 0  
    if Days_Overdue is None:
        Days_Overdue = 0  

    insert_query = "INSERT INTO TableData (Date, Chart_Type, Customer_Id, Customer_Name, Days_Overdue, Amount_Outstanding, Recovery) VALUES (%s, %s, %s, %s, %s, %s, %s);"
    cursor.execute(insert_query, (Date, Chart_Type, Customer_Id, Customer_Name, Days_Overdue, Amount_Outstanding, Recovery))

conn.commit()
cursor.close()
conn.close()
